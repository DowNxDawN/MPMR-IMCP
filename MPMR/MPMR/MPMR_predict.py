import os
import gc
import time
import numpy as np
import torch
import pandas as pd
import shutil
from PIL import Image
import matplotlib.pyplot as plt
from torchvision import transforms
import torch.nn as nn
from tqdm import tqdm
from openpyxl import load_workbook
import timm
from timm.layers import SwiGLUPacked
from peft import LoraConfig, get_peft_model
from torch.amp import autocast
from multiprocessing import Pool, cpu_count
import logging
from concurrent.futures import ThreadPoolExecutor

# 设置环境变量以优化资源利用
os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:512"
os.environ["OMP_NUM_THREADS"] = "4"
os.environ["MKL_NUM_THREADS"] = "4"
Image.MAX_IMAGE_PIXELS = None  # 允许处理大图像

# 配置日志记录
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


# 清理内存函数
def clean_memory():
    gc.collect()
    torch.cuda.empty_cache()


# 询问是否生成热图
generate_heatmap_flag = input("是否生成热图？(y/n): ").strip().lower() == "y"
use_multiprocessing = input("是否使用多进程进行图像分割？(y/n): ").strip().lower() == "y"

if use_multiprocessing:
    num_processes = int(input("请输入要使用的进程数量: ").strip())
else:
    num_processes = 1

# 自定义内容 - 需要修改为您的模型路径
model_paths = {
    "CD10": "/path/to/your/model.pth",
    "MUC2": "/path/to/your/model.pth",
    "MUC5AC": "/path/to/your/model.pth",
    "MUC6": "/path/to/your/model.pth",
}

# 获取用户输入的WSI图像目录
wsi_image_dir = "/path/to/your/data"
output_heatmap_dir = "/path/to/your/result/heatmaps"
output_excel_path = "/path/to/your/result/predict/predict.xlsx"

# 创建临时目录用于存储切分的图像块
temp_patches_dir = "/path/to/your/result/temp_patches"
os.makedirs(temp_patches_dir, exist_ok=True)

# 创建输出目录（如果不存在）
os.makedirs(output_heatmap_dir, exist_ok=True)
os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

# 图像处理参数
patch_size = 224
stride = 224  # 步长越小，预测越精细，但速度越慢
batch_size = 16  # 批量大小

# 背景图块筛选条件
mean_threshold_low = 143.32 - 2 * 41.30
mean_threshold_high = 143.32 + 2 * 41.30
std_threshold_low = 16.12 - 2 * 20.21
std_threshold_high = 16.12 + 2 * 20.21

# 设置设备
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"使用设备: {device}")

# 定义图像预处理
transform = transforms.Compose(
    [
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5]),
    ]
)


# 定义ViT-UNI2模型初始化函数
def initialize_model(model_path, num_classes=2):
    print(f"加载模型: {os.path.basename(model_path)}...")

    # 创建UNI2模型配置
    timm_kwargs = {
        "model_name": "vit_giant_patch14_224",
        "img_size": 224,
        "patch_size": 14,
        "depth": 24,
        "num_heads": 24,
        "init_values": 1e-5,
        "embed_dim": 1536,
        "mlp_ratio": 2.66667 * 2,
        "num_classes": 0,
        "no_embed_class": True,
        "mlp_layer": timm.layers.SwiGLUPacked,
        "act_layer": torch.nn.SiLU,
        "reg_tokens": 8,
        "dynamic_img_size": True,
        "in_chans": 3,
    }

    # 创建基础模型
    base_model = timm.create_model(pretrained=False, **timm_kwargs)

    # LoRA配置
    lora_config = LoraConfig(r=16, lora_alpha=16, target_modules=["qkv", "fc1", "fc2"], lora_dropout=0.05, bias="none")

    # 构建最终模型
    model = nn.Sequential(
        get_peft_model(base_model, lora_config), nn.Dropout(p=0.2), nn.Linear(base_model.num_features, num_classes)
    )

    # 加载训练好的权重
    state_dict = torch.load(model_path, map_location="cpu")

    # 处理可能的DataParallel模型保存格式
    if any(k.startswith("module.") for k in state_dict.keys()):
        from collections import OrderedDict

        new_state_dict = OrderedDict()
        for k, v in state_dict.items():
            name = k[7:] if k.startswith("module.") else k  # 去掉module前缀
            new_state_dict[name] = v
        state_dict = new_state_dict

    model.load_state_dict(state_dict)

    if torch.cuda.device_count() > 1:
        print(f"使用 {torch.cuda.device_count()} 个 GPU 进行预测")
        model = nn.DataParallel(model)

    model = model.to(device, memory_format=torch.channels_last)
    model.eval()

    return model


# 分割单个图像并保存图像块
def split_and_save_image(args):
    wsi_image_path, temp_dir, image_id = args
    logging.info(f"分割图像: {os.path.basename(wsi_image_path)}")

    try:
        # 加载WSI图片
        wsi_image = Image.open(wsi_image_path)

        # 获取WSI图片尺寸
        wsi_width, wsi_height = wsi_image.size
        image_name = os.path.basename(wsi_image_path).split(".")[0]

        # 为此图像创建临时目录
        image_temp_dir = os.path.join(temp_dir, image_name)
        os.makedirs(image_temp_dir, exist_ok=True)

        # 用于存储图像块位置信息的文件
        coords_file = os.path.join(image_temp_dir, "patch_coordinates.txt")

        patch_count = 0
        total_blocks = ((wsi_height - patch_size) // stride + 1) * ((wsi_width - patch_size) // stride + 1)

        with open(coords_file, "w") as f:
            # 记录图像尺寸信息
            f.write(f"width:{wsi_width},height:{wsi_height}\n")

            for y in tqdm(
                range(0, wsi_height - patch_size + 1, stride),
                desc=f"分割图像 {image_id}/{len(wsi_image_files)}",
                ncols=100,
            ):
                for x in range(0, wsi_width - patch_size + 1, stride):
                    # 裁剪图像块
                    patch = wsi_image.crop((x, y, x + patch_size, y + patch_size))

                    # 背景筛选
                    patch_np = np.array(patch.convert("L"))
                    patch_mean = patch_np.mean()
                    patch_std = patch_np.std()

                    # 跳过背景块
                    if not (
                        mean_threshold_low <= patch_mean <= mean_threshold_high
                        and std_threshold_low <= patch_std <= std_threshold_high
                    ):
                        # 保存图像块
                        patch_filename = f"patch_{y}_{x}.png"
                        patch_path = os.path.join(image_temp_dir, patch_filename)
                        patch.save(patch_path)

                        # 记录坐标
                        f.write(f"{patch_filename},{x},{y}\n")
                        patch_count += 1

        return image_name, patch_count, total_blocks

    except Exception as e:
        logging.error(f"处理图像 {os.path.basename(wsi_image_path)} 时发生错误: {str(e)}")
        return None, 0, 0


# 批量预测函数
def predict_batch(model, batch_tensor):
    """对批量图像块进行预测"""
    with torch.no_grad():
        with autocast(device_type="cuda", dtype=torch.float16):
            outputs = model(batch_tensor)
            probs = torch.softmax(outputs, dim=1)[:, 1].cpu().numpy()  # 获取正类概率
    return probs


# 获取目录下所有WSI图片文件
valid_extensions = (".jpg", ".jpeg", ".png", ".tif", ".tiff")
wsi_image_files = [
    os.path.join(wsi_image_dir, f) for f in os.listdir(wsi_image_dir) if f.lower().endswith(valid_extensions)
]

print(f"找到 {len(wsi_image_files)} 个图像文件")

# 创建一个DataFrame来存储结果
results_df = pd.DataFrame(
    columns=[
        "Image Name",
        "Total Patches",
        "MUC5AC Count",
        "MUC5AC Ratio",
        "MUC6 Count",
        "MUC6 Ratio",
        "MUC2 Count",
        "MUC2 Ratio",
        "CD10 Count",
        "CD10 Ratio",
        "Complete Type I Count",
        "Complete Type I Ratio",
        "Incomplete Type II Count",
        "Incomplete Type II Ratio",
        "Incomplete Type III Count",
        "Incomplete Type III Ratio",
        "Prediction Start Time",
        "Prediction Duration (s)",
    ]
)

# 如果Excel文件不存在，创建一个新的文件并写入列名
if not os.path.exists(output_excel_path):
    results_df.to_excel(output_excel_path, index=False)

# 第1阶段：分割图像并保存图像块
print("\n第1阶段: 分割图像并保存图像块...")

if use_multiprocessing and num_processes > 1:
    # 多进程分割图像
    with Pool(processes=num_processes) as pool:
        args_list = [(wsi_image_path, temp_patches_dir, i + 1) for i, wsi_image_path in enumerate(wsi_image_files)]
        results = list(pool.map(split_and_save_image, args_list))
else:
    # 单进程分割图像
    results = []
    for i, wsi_image_path in enumerate(wsi_image_files):
        result = split_and_save_image((wsi_image_path, temp_patches_dir, i + 1))
        results.append(result)

# 过滤掉None结果（表示处理失败的图像）
results = [r for r in results if r[0] is not None]
for image_name, patch_count, total_blocks in results:
    print(f"图像 {image_name} 处理完成: 保存了 {patch_count} 个非背景图像块，共 {total_blocks} 个潜在块")

# 第2阶段：加载模型并进行预测
print("\n第2阶段: 加载模型并进行预测...")

# 加载所有模型
models = {}
for key, path in model_paths.items():
    models[key] = initialize_model(path)
    print(f"模型 {key} 已加载")

# 主处理循环
for image_name, _, _ in results:
    print(f"\n正在预测图像: {image_name}")

    # 记录预测开始时间
    start_time = time.time()
    prediction_start_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(start_time))

    try:
        # 获取图像目录
        image_temp_dir = os.path.join(temp_patches_dir, image_name)
        coords_file = os.path.join(image_temp_dir, "patch_coordinates.txt")

        # 读取坐标信息
        coords_data = {}
        with open(coords_file, "r") as f:
            lines = f.readlines()
            # 第一行包含图像尺寸信息
            header = lines[0].strip().split(",")
            width = int(header[0].split(":")[1])
            height = int(header[1].split(":")[1])

            # 读取剩余行的坐标信息
            for line in lines[1:]:
                filename, x, y = line.strip().split(",")
                coords_data[filename] = (int(x), int(y))

        # 预测结果存储
        if generate_heatmap_flag:
            heatmaps = {key: np.zeros((height, width)) for key in model_paths.keys()}

        # 统计块的数量
        counts = {key: 0 for key in model_paths.keys()}
        total_patches = len(coords_data)

        # 额外结果存储
        incomplete_type_III_count = 0
        incomplete_type_II_count = 0
        complete_type_I_count = 0

        print(f"预计处理 {total_patches} 个图像块")

        # 获取所有图像块文件
        patch_files = list(coords_data.keys())

        # 批量处理图像块
        for i in tqdm(range(0, len(patch_files), batch_size), desc="批量预测", ncols=100):
            batch_files = patch_files[i : i + batch_size]

            patches = []
            batch_coords = []

            for filename in batch_files:
                patch_path = os.path.join(image_temp_dir, filename)
                patch = Image.open(patch_path)
                patches.append(transform(patch).unsqueeze(0))
                batch_coords.append(coords_data[filename])

            # 合并批次
            if patches:
                patches_tensor = torch.cat(patches).to(device, memory_format=torch.channels_last)

                # 为每个模型进行预测
                results = {}
                for key, model in models.items():
                    probs = predict_batch(model, patches_tensor)
                    results[key] = probs

                # 处理预测结果
                for i, (x, y) in enumerate(batch_coords):
                    # 模型预测结果
                    cd10_prob = 1 - results["CD10"][i]
                    muc2_prob = 1 - results["MUC2"][i]
                    muc5ac_prob = 1 - results["MUC5AC"][i]
                    muc6_prob = 1 - results["MUC6"][i]

                    # 更新热图
                    if generate_heatmap_flag:
                        for key in model_paths.keys():
                            prob = 1 - results[key][i]
                            heatmaps[key][y : y + patch_size, x : x + patch_size] = prob

                    # 统计阳性块数
                    for key, prob in zip(models.keys(), [cd10_prob, muc2_prob, muc5ac_prob, muc6_prob]):
                        if prob >= 0.5:
                            counts[key] += 1

                    # 肠化类型判断
                    if cd10_prob >= 0.5 and muc2_prob >= 0.5 and muc5ac_prob < 0.5:
                        incomplete_type_III_count += 1
                    elif muc2_prob >= 0.5 and muc5ac_prob < 0.5:
                        incomplete_type_II_count += 1
                    elif muc2_prob < 0.5 and cd10_prob < 0.5 and muc5ac_prob >= 0.5 and muc6_prob >= 0.5:
                        complete_type_I_count += 1

        # 记录预测结束时间并计算耗时
        end_time = time.time()
        prediction_duration = end_time - start_time

        # 计算非负类别图块的总数
        non_negative_total = sum(counts.values())

        # 计算块的比例
        if non_negative_total > 0:
            ratios = {key: counts[key] / non_negative_total for key in model_paths.keys()}
        else:
            ratios = {key: 0 for key in model_paths.keys()}

        # 计算肠化类型的总数和比例
        intestinal_metaplasia_total = incomplete_type_III_count + incomplete_type_II_count + complete_type_I_count

        if intestinal_metaplasia_total > 0:
            incomplete_type_III_ratio = incomplete_type_III_count / intestinal_metaplasia_total
            incomplete_type_II_ratio = incomplete_type_II_count / intestinal_metaplasia_total
            complete_type_I_ratio = complete_type_I_count / intestinal_metaplasia_total
        else:
            incomplete_type_III_ratio = 0
            incomplete_type_II_ratio = 0
            complete_type_I_ratio = 0

        # 将结果添加到DataFrame
        new_row = pd.DataFrame(
            {
                "Image Name": [image_name],
                "Total Patches": [total_patches],
                "MUC5AC Count": [counts["MUC5AC"]],
                "MUC5AC Ratio": [ratios["MUC5AC"]],
                "MUC6 Count": [counts["MUC6"]],
                "MUC6 Ratio": [ratios["MUC6"]],
                "MUC2 Count": [counts["MUC2"]],
                "MUC2 Ratio": [ratios["MUC2"]],
                "CD10 Count": [counts["CD10"]],
                "CD10 Ratio": [ratios["CD10"]],
                "Complete Type I Count": [complete_type_I_count],
                "Complete Type I Ratio": [complete_type_I_ratio],
                "Incomplete Type II Count": [incomplete_type_II_count],
                "Incomplete Type II Ratio": [incomplete_type_II_ratio],
                "Incomplete Type III Count": [incomplete_type_III_count],
                "Incomplete Type III Ratio": [incomplete_type_III_ratio],
                "Prediction Start Time": [prediction_start_time],
                "Prediction Duration (s)": [prediction_duration],
            }
        )

        # 追加结果到Excel文件
        try:
            # 尝试追加到现有文件
            book = load_workbook(output_excel_path)
            writer = pd.ExcelWriter(output_excel_path, engine="openpyxl")
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            new_row.to_excel(
                writer, sheet_name="Sheet1", index=False, header=False, startrow=writer.sheets["Sheet1"].max_row
            )
            writer.close()
        except:
            # 如果出错，创建新文件
            if not os.path.exists(output_excel_path):
                new_row.to_excel(output_excel_path, index=False)
            else:
                # 读取现有Excel并追加
                existing_df = pd.read_excel(output_excel_path)
                combined_df = pd.concat([existing_df, new_row], ignore_index=True)
                combined_df.to_excel(output_excel_path, index=False)

        # 生成热图
        if generate_heatmap_flag:
            # 重新加载原始图像用于背景
            wsi_image_path = None
            for path in wsi_image_files:
                if os.path.basename(path).split(".")[0] == image_name:
                    wsi_image_path = path
                    break

            if wsi_image_path:
                wsi_image = Image.open(wsi_image_path)

                for key in model_paths.keys():
                    plt.figure(figsize=(12, 10))
                    plt.imshow(wsi_image, cmap="gray", alpha=0.7)
                    plt.imshow(heatmaps[key], cmap="jet", alpha=0.5, vmin=0, vmax=1)
                    plt.colorbar(label="Positive Probability")
                    plt.title(f"WSI Heatmap - {key}")

                    output_heatmap_path = os.path.join(output_heatmap_dir, f"heatmap_{image_name}_ViT-UNI2-{key}.png")
                    plt.savefig(output_heatmap_path, dpi=300, bbox_inches="tight")
                    plt.close()
                    print(f"已保存热图: {output_heatmap_path}")

        # 打印统计结果
        print(f"\n图像 {image_name} 预测完成")
        print(f"总计处理块数: {total_patches}")
        print(f"MUC5AC 阳性块数: {counts['MUC5AC']} ({ratios['MUC5AC']:.2%})")
        print(f"MUC6 阳性块数: {counts['MUC6']} ({ratios['MUC6']:.2%})")
        print(f"MUC2 阳性块数: {counts['MUC2']} ({ratios['MUC2']:.2%})")
        print(f"CD10 阳性块数: {counts['CD10']} ({ratios['CD10']:.2%})")
        print(f"完全型肠化I型: {complete_type_I_count} ({complete_type_I_ratio:.2%})")
        print(f"不完全型肠化II型: {incomplete_type_II_count} ({incomplete_type_II_ratio:.2%})")
        print(f"不完全型肠化III型: {incomplete_type_III_count} ({incomplete_type_III_ratio:.2%})")
        print(f"处理耗时: {prediction_duration:.2f}秒")

    except Exception as e:
        print(f"处理图像 {image_name} 时发生错误: {str(e)}")
        # 记录错误到Excel
        error_row = pd.DataFrame(
            {
                "Image Name": [image_name],
                "Total Patches": [0],
                "MUC5AC Count": [0],
                "MUC5AC Ratio": [0],
                "MUC6 Count": [0],
                "MUC6 Ratio": [0],
                "MUC2 Count": [0],
                "MUC2 Ratio": [0],
                "CD10 Count": [0],
                "CD10 Ratio": [0],
                "Complete Type I Count": [0],
                "Complete Type I Ratio": [0],
                "Incomplete Type II Count": [0],
                "Incomplete Type II Ratio": [0],
                "Incomplete Type III Count": [0],
                "Incomplete Type III Ratio": [0],
                "Prediction Start Time": [prediction_start_time],
                "Prediction Duration (s)": [0],
                "Error": [str(e)],
            }
        )

        if os.path.exists(output_excel_path):
            existing_df = pd.read_excel(output_excel_path)
            combined_df = pd.concat([existing_df, error_row], ignore_index=True)
            combined_df.to_excel(output_excel_path, index=False)
        else:
            error_row.to_excel(output_excel_path, index=False)

# 询问是否清理临时文件
keep_temp = input("\n是否保留临时分割图像？(y/n): ").strip().lower() == "y"
if not keep_temp:
    print("正在清理临时文件...")
    shutil.rmtree(temp_patches_dir)
    print("临时文件已删除")

print("\n所有图像处理完成!")
print(f"结果已保存到: {output_excel_path}")
if generate_heatmap_flag:
    print(f"热图已保存到: {output_heatmap_dir}")
